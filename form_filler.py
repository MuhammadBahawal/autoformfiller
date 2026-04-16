from __future__ import annotations

import argparse
import json
import math
import random
import re
import threading
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable

import requests
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


TRACKING_DEFAULTS = {
    "status_column": "__status",
    "message_column": "__message",
    "processed_at_column": "__processed_at",
    "profile_column": "__profile_id",
}

SURVEY_DEFAULT_OPTION_PRIORITY = (
    "none of the above",
    "none above",
    "no above",
    "never",
    "no",
    "none",
    "not applicable",
)
SURVEY_FINAL_CTA_KEYWORDS = (
    "get",
    "claim",
    "view",
    "see",
    "start",
    "finish",
    "complete",
    "done",
)
SURVEY_CONTINUE_KEYWORDS = ("continue", "next", "proceed")
FORM_INITIAL_STEP_DEFAULT_COLUMNS = ("zip", "zip code", "zipcode", "postal", "postal code")
FORM_INITIAL_STEP_DEFAULT_ACTION_KEYWORDS = SURVEY_CONTINUE_KEYWORDS
SURVEY_NAVIGATION_KEYWORDS = (
    "continue",
    "next",
    "proceed",
    "back",
    "previous",
    "submit",
    *SURVEY_FINAL_CTA_KEYWORDS,
)
SURVEY_SPECIAL_QUESTION_RULES = (
    (
        (
            "what is your current employment status",
            "current employment status",
            "employment status",
        ),
        ("employed",),
    ),
    (
        (
            "do you own an active bank account",
            "own an active bank account",
            "active bank account",
        ),
        ("yes",),
    ),
)
SURVEY_COMPLETION_PAGE_KEYWORDS = (
    "thank you",
    "all set",
    "you are done",
    "survey completed",
    "submission received",
)
SURVEY_OFFER_PAGE_KEYWORDS = (
    "deal below",
    "show more offers",
    "ways to earn",
    "earn & save",
    "special offers",
)
SURVEY_OFFER_ACTION_PRIORITY = (
    "get",
    "claim",
    "check eligibility",
    "check",
    "continue",
    "see affordable options",
    "see",
    "view",
    "start",
)
SURVEY_ACTION_EXCLUDE_PHRASES = (
    "show more offers",
    "back",
    "previous",
)


@dataclass
class RowTask:
    row_number: int
    values: dict[str, str]


@dataclass
class SurveyOption:
    kind: str
    text: str
    normalized_text: str
    element: Any
    control: Any | None = None
    group_key: str = ""
    question_text: str = ""


@dataclass
class FormRuntimeState:
    last_form_url: str = ""
    last_form_handle: str = ""


@dataclass
class SurveyPageContext:
    handle: str
    url: str
    page_text: str
    radio_options: list[SurveyOption]
    checkbox_options: list[SurveyOption]
    button_options: list[SurveyOption]
    actions: list[tuple[Any, str, str]]

    @property
    def survey_options(self) -> list[SurveyOption]:
        options = self.radio_options + self.checkbox_options
        if options:
            return options
        return self.button_options


def log(message: str) -> None:
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] {message}", flush=True)


def normalize_key(value: str) -> str:
    collapsed = re.sub(r"\s+", " ", str(value or "").strip())
    return collapsed.lower()


def cell_to_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_selector(raw_selector: str) -> tuple[str, str]:
    selector = raw_selector.strip()
    selector_map = {
        "css=": By.CSS_SELECTOR,
        "xpath=": By.XPATH,
        "id=": By.ID,
        "name=": By.NAME,
    }
    for prefix, by in selector_map.items():
        if selector.startswith(prefix):
            return by, selector[len(prefix) :]
    if selector.startswith("//") or selector.startswith("(//"):
        return By.XPATH, selector
    return By.CSS_SELECTOR, selector


def load_config(config_path: Path) -> dict[str, Any]:
    if not config_path.exists():
        raise FileNotFoundError(
            f"Config file not found: {config_path}. Copy config.example.json to config.json and edit it first."
        )
    return json.loads(config_path.read_text(encoding="utf-8"))


def safe_save_workbook(workbook, path: Path) -> None:
    try:
        workbook.save(path)
    except PermissionError as exc:
        raise PermissionError(
            f"Cannot save Excel file '{path}'. Please close the Excel file and run the script again."
        ) from exc


class ExcelTracker:
    def __init__(self, excel_config: dict[str, Any]) -> None:
        self.path = Path(excel_config["path"]).expanduser().resolve()
        if not self.path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.path}")

        self.header_row = int(excel_config.get("header_row", 1))
        self.first_data_row = int(excel_config.get("first_data_row", self.header_row + 1))
        self.sheet_name = excel_config.get("sheet_name")

        tracking = TRACKING_DEFAULTS | {
            key: excel_config.get(key, default)
            for key, default in TRACKING_DEFAULTS.items()
        }
        self.status_column_name = tracking["status_column"]
        self.message_column_name = tracking["message_column"]
        self.processed_at_column_name = tracking["processed_at_column"]
        self.profile_column_name = tracking["profile_column"]

        self.lock = threading.Lock()
        self.workbook = load_workbook(self.path)
        self.sheet = self.workbook[self.sheet_name] if self.sheet_name else self.workbook.active
        self.max_column = self.sheet.max_column
        self._header_index_by_normalized: dict[str, int] = {}
        self._header_index_by_exact: dict[str, int] = {}
        self._refresh_headers()
        self._ensure_tracking_columns()
        self._refresh_headers()
        self.source_header_names = self._build_source_header_names()

    def _refresh_headers(self) -> None:
        header_index_by_normalized: dict[str, int] = {}
        header_index_by_exact: dict[str, int] = {}
        max_col = self.sheet.max_column
        for col_idx in range(1, max_col + 1):
            header_value = self.sheet.cell(row=self.header_row, column=col_idx).value
            if header_value is None:
                continue
            exact = str(header_value).strip()
            if not exact:
                continue
            normalized = normalize_key(exact)
            if normalized in header_index_by_normalized:
                raise ValueError(
                    f"Duplicate Excel header detected after normalization: '{exact}'. Please rename duplicate columns."
                )
            header_index_by_normalized[normalized] = col_idx
            header_index_by_exact[exact] = col_idx
        self._header_index_by_normalized = header_index_by_normalized
        self._header_index_by_exact = header_index_by_exact

    def _ensure_tracking_columns(self) -> None:
        for column_name in (
            self.status_column_name,
            self.message_column_name,
            self.processed_at_column_name,
            self.profile_column_name,
        ):
            normalized = normalize_key(column_name)
            if normalized not in self._header_index_by_normalized:
                new_col_idx = self.sheet.max_column + 1
                self.sheet.cell(row=self.header_row, column=new_col_idx, value=column_name)
        safe_save_workbook(self.workbook, self.path)

    def _build_source_header_names(self) -> list[str]:
        tracking_names = {
            normalize_key(self.status_column_name),
            normalize_key(self.message_column_name),
            normalize_key(self.processed_at_column_name),
            normalize_key(self.profile_column_name),
        }
        headers: list[str] = []
        for exact_name in self._header_index_by_exact:
            if normalize_key(exact_name) not in tracking_names:
                headers.append(exact_name)
        return headers

    @property
    def available_headers(self) -> list[str]:
        return list(self.source_header_names)

    def build_pending_tasks(self) -> list[RowTask]:
        status_col_idx = self._header_index_by_normalized[normalize_key(self.status_column_name)]
        tasks: list[RowTask] = []
        for row_number in range(self.first_data_row, self.sheet.max_row + 1):
            row_values: dict[str, str] = {}
            is_empty = True
            for col_idx in range(1, self.sheet.max_column + 1):
                raw_value = self.sheet.cell(row=row_number, column=col_idx).value
                text_value = cell_to_string(raw_value)
                column_letter = self.sheet.cell(row=self.header_row, column=col_idx).column_letter.lower()
                row_values[f"__col_{column_letter}"] = text_value
                header_value = self.sheet.cell(row=self.header_row, column=col_idx).value
                if header_value is not None and str(header_value).strip():
                    row_values[normalize_key(str(header_value))] = text_value
                if text_value:
                    is_empty = False
            if is_empty:
                continue
            status_value = cell_to_string(self.sheet.cell(row=row_number, column=status_col_idx).value).upper()
            if status_value == "DONE":
                continue
            tasks.append(RowTask(row_number=row_number, values=row_values))
        return tasks

    def assert_column_exists(self, column_name: str) -> None:
        normalized = normalize_key(column_name)
        if normalized in self._header_index_by_normalized:
            return
        available = ", ".join(self.available_headers)
        raise ValueError(f"Excel column missing: {column_name}. Available columns: {available}")

    def assert_column_letter_exists(self, column_letter: str) -> None:
        normalized_letter = str(column_letter).strip().upper()
        if not normalized_letter or not normalized_letter.isalpha():
            raise ValueError(f"Invalid Excel column letter: {column_letter}")
        col_number = 0
        for char in normalized_letter:
            col_number = col_number * 26 + (ord(char) - 64)
        if col_number < 1 or col_number > self.sheet.max_column:
            raise ValueError(
                f"Excel column letter {column_letter} is outside the current sheet range (max column {self.sheet.max_column})."
            )

    def assert_field_sources_exist(self, fields: list[dict[str, Any]]) -> None:
        for field in fields:
            column_letter = str(field.get("excel_column_letter", "")).strip()
            if column_letter:
                self.assert_column_letter_exists(column_letter)
                continue
            self.assert_column_exists(field["column"])

    def assert_columns_exist(self, required_columns: list[str]) -> None:
        missing = [column for column in required_columns if normalize_key(column) not in self._header_index_by_normalized]
        if missing:
            available = ", ".join(self.available_headers)
            missing_text = ", ".join(missing)
            raise ValueError(f"Excel columns missing: {missing_text}. Available columns: {available}")

    def mark_result(self, row_number: int, status: str, message: str, profile_id: str) -> None:
        with self.lock:
            status_col = self._header_index_by_normalized[normalize_key(self.status_column_name)]
            message_col = self._header_index_by_normalized[normalize_key(self.message_column_name)]
            processed_at_col = self._header_index_by_normalized[normalize_key(self.processed_at_column_name)]
            profile_col = self._header_index_by_normalized[normalize_key(self.profile_column_name)]

            self.sheet.cell(row=row_number, column=status_col, value=status)
            self.sheet.cell(row=row_number, column=message_col, value=message[:30000])
            self.sheet.cell(
                row=row_number,
                column=processed_at_col,
                value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            )
            self.sheet.cell(row=row_number, column=profile_col, value=profile_id)
            safe_save_workbook(self.workbook, self.path)


class AdsPowerClient:
    def __init__(self, adspower_config: dict[str, Any]) -> None:
        self.base_url = str(adspower_config.get("base_url", "http://local.adspower.net:50325")).rstrip("/")
        self.api_key = str(adspower_config.get("api_key", "")).strip()
        self.keep_browser_open = bool(adspower_config.get("keep_browser_open", False))

    def _headers(self) -> dict[str, str]:
        if not self.api_key:
            return {}
        return {"Authorization": f"Bearer {self.api_key}"}

    def _get_json(self, path: str, timeout: int = 60) -> dict[str, Any]:
        response = requests.get(f"{self.base_url}{path}", headers=self._headers(), timeout=timeout)
        response.raise_for_status()
        return response.json()

    def _connect_driver(self, browser_data: dict[str, Any]) -> webdriver.Chrome:
        chrome_driver_path = browser_data["webdriver"]
        debugger_address = browser_data["ws"]["selenium"]

        options = Options()
        options.add_experimental_option("debuggerAddress", debugger_address)
        service = Service(executable_path=chrome_driver_path)
        return webdriver.Chrome(service=service, options=options)

    def list_active_profiles(self) -> list[dict[str, Any]]:
        payload = self._get_json("/api/v1/browser/local-active", timeout=30)
        if payload.get("code") != 0:
            raise RuntimeError(f"AdsPower local-active failed: {payload.get('msg', 'unknown error')}")
        return payload.get("data", {}).get("list", [])

    def start_profile(self, profile_id: str) -> tuple[webdriver.Chrome, dict[str, Any]]:
        payload = self._get_json(f"/api/v1/browser/start?user_id={profile_id}", timeout=60)
        if payload.get("code") != 0:
            raise RuntimeError(f"AdsPower start failed for {profile_id}: {payload.get('msg', 'unknown error')}")
        driver = self._connect_driver(payload["data"])
        return driver, payload

    def attach_to_active_profile(self, profile_data: dict[str, Any]) -> webdriver.Chrome:
        return self._connect_driver(profile_data)

    def stop_profile(self, profile_id: str) -> None:
        if self.keep_browser_open:
            return
        response = requests.get(
            f"{self.base_url}/api/v1/browser/stop?user_id={profile_id}",
            headers=self._headers(),
            timeout=30,
        )
        response.raise_for_status()


def get_row_value(row_values: dict[str, str], column_name: str) -> str:
    return row_values.get(normalize_key(column_name), "")


def get_field_value(row_values: dict[str, str], field_config: dict[str, Any]) -> str:
    column_letter = str(field_config.get("excel_column_letter", "")).strip().lower()
    if column_letter:
        return row_values.get(f"__col_{column_letter}", "")
    return get_row_value(row_values, field_config["column"])


def parse_date_value(value: str) -> datetime:
    cleaned = value.strip()
    if not cleaned:
        raise ValueError("Date value is empty")

    date_formats = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%m-%d-%Y",
        "%m-%d-%y",
    )
    for fmt in date_formats:
        try:
            return datetime.strptime(cleaned, fmt)
        except ValueError:
            continue
    raise ValueError(f"Unsupported date format: {value}")


def transform_value(raw_value: str, field_config: dict[str, Any]) -> str:
    transform = str(field_config.get("transform", "")).strip().lower()
    if not transform:
        return raw_value

    if transform in {"date_month_number", "date_day_number", "date_year"}:
        parsed = parse_date_value(raw_value)
        if transform == "date_month_number":
            return str(parsed.month)
        if transform == "date_day_number":
            return str(parsed.day)
        return str(parsed.year)

    raise ValueError(f"Unsupported transform '{transform}'")


def text_contains_phrase(text: str, phrase: str) -> bool:
    normalized_text = normalize_key(text)
    normalized_phrase = normalize_key(phrase)
    if not normalized_text or not normalized_phrase:
        return False
    pattern = rf"\b{re.escape(normalized_phrase)}\b"
    return re.search(pattern, normalized_text) is not None


def matches_any_phrase(text: str, phrases: tuple[str, ...]) -> bool:
    return any(text_contains_phrase(text, phrase) for phrase in phrases)


def find_first_element(
    driver: webdriver.Chrome,
    selectors: list[str],
    timeout_seconds: int,
    require_displayed: bool = True,
):
    deadline = time.time() + timeout_seconds
    last_error = None
    while time.time() < deadline:
        for selector in selectors:
            by, query = parse_selector(selector)
            try:
                matches = driver.find_elements(by, query)
            except WebDriverException as exc:
                last_error = exc
                continue
            for element in matches:
                try:
                    if not require_displayed or element.is_displayed():
                        return element
                except WebDriverException as exc:
                    last_error = exc
        time.sleep(0.25)
    selector_text = " | ".join(selectors)
    if last_error:
        raise TimeoutException(f"Element not found for selectors: {selector_text}. Last error: {last_error}") from last_error
    raise TimeoutException(f"Element not found for selectors: {selector_text}")


def try_find_first_element(
    driver: webdriver.Chrome,
    selectors: list[str],
    require_displayed: bool = True,
):
    for selector in selectors:
        by, query = parse_selector(selector)
        try:
            matches = driver.find_elements(by, query)
        except WebDriverException:
            continue
        for element in matches:
            try:
                if not require_displayed or element.is_displayed():
                    return element
            except WebDriverException:
                continue
    return None


def get_element_text(driver: webdriver.Chrome, element) -> str:
    try:
        value = driver.execute_script(
            """
            const element = arguments[0];
            const raw =
                element.innerText ||
                element.textContent ||
                element.value ||
                element.getAttribute('aria-label') ||
                element.getAttribute('title') ||
                '';
            return raw.replace(/\\s+/g, ' ').trim();
            """,
            element,
        )
    except WebDriverException:
        return ""
    return str(value or "").strip()


def get_page_text(driver: webdriver.Chrome) -> str:
    try:
        value = driver.execute_script(
            """
            const body = document.body;
            if (!body) {
                return '';
            }
            return (body.innerText || body.textContent || '').replace(/\\s+/g, ' ').trim();
            """
        )
    except WebDriverException:
        return ""
    return normalize_key(str(value or ""))


def click_element(driver: webdriver.Chrome, element) -> None:
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
    try:
        element.click()
    except WebDriverException:
        driver.execute_script("arguments[0].click();", element)


def get_field_selectors(field_config: dict[str, Any]) -> list[str]:
    selectors = [str(item).strip() for item in field_config.get("selectors") or [] if str(item).strip()]
    legacy_selector = str(field_config.get("selector", "")).strip()
    if legacy_selector and legacy_selector not in selectors:
        selectors.insert(0, legacy_selector)
    return selectors


def field_is_visible_now(driver: webdriver.Chrome, field_config: dict[str, Any]) -> bool:
    selectors = get_field_selectors(field_config)
    if not selectors:
        return False
    return try_find_first_element(driver, selectors, require_displayed=True) is not None


def wait_for_any_field_visible(
    driver: webdriver.Chrome,
    field_configs: list[dict[str, Any]],
    timeout_seconds: float,
) -> bool:
    if timeout_seconds <= 0:
        return any(field_is_visible_now(driver, field_config) for field_config in field_configs)

    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        if any(field_is_visible_now(driver, field_config) for field_config in field_configs):
            return True
        time.sleep(0.25)
    return any(field_is_visible_now(driver, field_config) for field_config in field_configs)


def get_visible_field_element(driver: webdriver.Chrome, field_config: dict[str, Any]):
    selectors = get_field_selectors(field_config)
    if not selectors:
        return None
    return try_find_first_element(driver, selectors, require_displayed=True)


def blur_element(driver: webdriver.Chrome, element) -> None:
    try:
        driver.execute_script(
            """
            const element = arguments[0];
            if (!element) {
                return;
            }
            element.dispatchEvent(new Event('blur', { bubbles: true }));
            if (typeof element.blur === 'function') {
                element.blur();
            }
            """,
            element,
        )
    except WebDriverException:
        return


def page_has_configured_field(driver: webdriver.Chrome, fields: list[dict[str, Any]]) -> bool:
    for field_config in fields:
        if str(field_config.get("type", "text")).lower() == "hidden_text":
            continue
        if field_is_visible_now(driver, field_config):
            return True
    return False


def get_choice_click_target(driver: webdriver.Chrome, choice_input):
    try:
        return driver.execute_script(
            """
            const input = arguments[0];
            if (input.id) {
                for (const label of document.querySelectorAll('label')) {
                    if (label.htmlFor === input.id) {
                        return label;
                    }
                }
            }
            const closestLabel = input.closest('label');
            if (closestLabel) {
                return closestLabel;
            }
            return input;
            """,
            choice_input,
        )
    except WebDriverException:
        return choice_input


def get_choice_text_candidates(driver: webdriver.Chrome, choice_input) -> list[str]:
    try:
        values = driver.execute_script(
            """
            const input = arguments[0];
            const results = [];
            const seen = new Set();
            function push(text) {
                const cleaned = (text || '').replace(/\\s+/g, ' ').trim();
                if (!cleaned) {
                    return;
                }
                const key = cleaned.toLowerCase();
                if (!seen.has(key)) {
                    seen.add(key);
                    results.push(cleaned);
                }
            }

            push(input.getAttribute('aria-label'));
            push(input.getAttribute('title'));

            if (input.id) {
                for (const label of document.querySelectorAll('label')) {
                    if (label.htmlFor === input.id) {
                        push(label.innerText || label.textContent);
                    }
                }
            }

            const closestLabel = input.closest('label');
            if (closestLabel) {
                push(closestLabel.innerText || closestLabel.textContent);
            }

            let sibling = input.nextElementSibling;
            while (sibling && results.length < 5) {
                push(sibling.innerText || sibling.textContent);
                sibling = sibling.nextElementSibling;
            }

            if (input.parentElement) {
                push(input.parentElement.innerText || input.parentElement.textContent);
            }

            return results;
            """,
            choice_input,
        )
    except WebDriverException:
        return []
    return [str(value).strip() for value in values or [] if str(value).strip()]


def get_choice_question_text(driver: webdriver.Chrome, choice_input) -> str:
    try:
        value = driver.execute_script(
            """
            const input = arguments[0];
            function clean(text) {
                return (text || '').replace(/\s+/g, ' ').trim();
            }
            function text(node) {
                if (!node) {
                    return '';
                }
                return clean(node.innerText || node.textContent || '');
            }
            function textFromIds(rawIds) {
                const ids = clean(rawIds).split(/\s+/).filter(Boolean);
                const values = [];
                for (const id of ids) {
                    const element = document.getElementById(id);
                    const value = text(element);
                    if (value) {
                        values.push(value);
                    }
                }
                return clean(values.join(' '));
            }

            let question = textFromIds(input.getAttribute('aria-labelledby'));
            if (!question) {
                const fieldset = input.closest('fieldset');
                if (fieldset) {
                    question = text(fieldset.querySelector('legend'));
                }
            }
            if (!question) {
                const group = input.closest("[role='radiogroup'], [role='group']");
                if (group) {
                    question =
                        text(group.querySelector('legend, [role="heading"], h1, h2, h3, h4, h5, h6, label, p')) ||
                        text(group);
                }
            }
            if (!question) {
                const section = input.closest('[data-question], [data-testid*="question"], .question, .survey-question');
                if (section) {
                    question =
                        text(section.querySelector('legend, [role="heading"], h1, h2, h3, h4, h5, h6, label, p')) ||
                        text(section);
                }
            }
            return question;
            """,
            choice_input,
        )
    except WebDriverException:
        return ""
    return str(value or "").strip()


def get_choice_container_key(driver: webdriver.Chrome, choice_input) -> str:
    try:
        value = driver.execute_script(
            """
            const input = arguments[0];
            function clean(text) {
                return (text || '').replace(/\s+/g, ' ').trim();
            }
            function classKey(node) {
                const raw = typeof node.className === 'string' ? node.className : '';
                return raw.split(/\s+/).filter(Boolean).slice(0, 4).join('.');
            }
            function keyFor(node) {
                if (!node) {
                    return '';
                }
                return clean(
                    [
                        node.tagName ? node.tagName.toLowerCase() : '',
                        node.id || '',
                        node.getAttribute('name') || '',
                        node.getAttribute('data-question') || '',
                        node.getAttribute('data-testid') || '',
                        classKey(node),
                    ]
                        .filter(Boolean)
                        .join('|')
                );
            }

            const selector = "input[type='" + input.type + "']";
            let node = input.parentElement;
            while (node) {
                try {
                    const count = node.querySelectorAll(selector).length;
                    const text = clean(node.innerText || node.textContent || '');
                    if (count > 1 && text && text.length <= 300) {
                        return keyFor(node);
                    }
                } catch (error) {
                }
                node = node.parentElement;
            }
            return '';
            """,
            choice_input,
        )
    except WebDriverException:
        return ""
    return normalize_key(str(value or ""))


def get_choice_group_key(driver: webdriver.Chrome, choice_input, question_text: str, input_type: str) -> str:
    normalized_question = normalize_key(question_text)
    container_key = get_choice_container_key(driver, choice_input)
    if input_type == "checkbox":
        if normalized_question:
            return normalized_question
        if container_key:
            return container_key

    try:
        value = driver.execute_script(
            """
            const input = arguments[0];
            return (
                input.name ||
                input.getAttribute('data-question') ||
                input.getAttribute('data-testid') ||
                input.id ||
                input.value ||
                ''
            );
            """,
            choice_input,
        )
    except WebDriverException:
        value = ""

    normalized = normalize_key(str(value or ""))
    if normalized:
        return normalized

    if normalized_question:
        return normalized_question
    if container_key:
        return container_key
    return ""


def get_choice_text(driver: webdriver.Chrome, choice_input) -> str:
    for candidate in get_choice_text_candidates(driver, choice_input):
        normalized = normalize_key(candidate)
        if normalized and len(normalized) <= 120:
            return candidate
    target = get_choice_click_target(driver, choice_input)
    return get_element_text(driver, target)


def collect_input_options(driver: webdriver.Chrome, input_type: str) -> list[SurveyOption]:
    options: list[SurveyOption] = []
    for choice_input in driver.find_elements(By.CSS_SELECTOR, f"input[type='{input_type}']"):
        click_target = get_choice_click_target(driver, choice_input)
        try:
            is_visible = choice_input.is_displayed() or click_target.is_displayed()
        except WebDriverException:
            is_visible = False
        if not is_visible:
            continue
        text = get_choice_text(driver, choice_input)
        normalized_text = normalize_key(text)
        if not normalized_text:
            continue
        question_text = get_choice_question_text(driver, choice_input)
        group_key = get_choice_group_key(driver, choice_input, question_text, input_type)
        options.append(
            SurveyOption(
                kind=input_type,
                text=text,
                normalized_text=normalized_text,
                element=click_target,
                control=choice_input,
                group_key=group_key,
                question_text=question_text,
            )
        )
    return options


def collect_checkbox_options(driver: webdriver.Chrome) -> list[SurveyOption]:
    return collect_input_options(driver, "checkbox")


def collect_radio_options(driver: webdriver.Chrome) -> list[SurveyOption]:
    return collect_input_options(driver, "radio")


def collect_button_options(driver: webdriver.Chrome) -> list[SurveyOption]:
    options: list[SurveyOption] = []
    selector = "button, [role='button'], input[type='button'], input[type='submit'], a"
    for element in driver.find_elements(By.CSS_SELECTOR, selector):
        try:
            if not element.is_displayed() or not element.is_enabled():
                continue
        except WebDriverException:
            continue

        text = get_element_text(driver, element)
        normalized_text = normalize_key(text)
        if not normalized_text:
            continue
        if matches_any_phrase(normalized_text, SURVEY_NAVIGATION_KEYWORDS):
            continue
        options.append(
            SurveyOption(
                kind="button",
                text=text,
                normalized_text=normalized_text,
                element=element,
            )
        )
    return options


def collect_clickable_actions(driver: webdriver.Chrome) -> list[tuple[Any, str, str]]:
    actions: list[tuple[Any, str, str]] = []
    selector = "button, [role='button'], input[type='button'], input[type='submit'], a"
    for element in driver.find_elements(By.CSS_SELECTOR, selector):
        try:
            if not element.is_displayed() or not element.is_enabled():
                continue
        except WebDriverException:
            continue
        text = get_element_text(driver, element)
        normalized_text = normalize_key(text)
        if not normalized_text:
            continue
        actions.append((element, text, normalized_text))
    return actions


def find_action_by_keywords(
    actions: list[tuple[Any, str, str]],
    keywords: tuple[str, ...],
) -> tuple[Any, str, str] | None:
    for keyword in keywords:
        for action in actions:
            _, _, normalized_text = action
            if normalized_text == keyword or normalized_text.startswith(f"{keyword} "):
                return action

    for keyword in keywords:
        for action in actions:
            _, _, normalized_text = action
            if text_contains_phrase(normalized_text, keyword):
                return action

    return None


def element_is_enabled(driver: webdriver.Chrome, element) -> bool:
    try:
        enabled = element.is_enabled()
    except WebDriverException:
        enabled = True

    try:
        script_enabled = driver.execute_script(
            """
            const element = arguments[0];
            if (!element) {
                return false;
            }
            const disabledAttr = element.getAttribute('disabled');
            const ariaDisabled = (element.getAttribute('aria-disabled') || '').toLowerCase();
            const computed = window.getComputedStyle(element);
            const pointerEvents = computed ? computed.pointerEvents : '';
            return !disabledAttr && ariaDisabled !== 'true' && pointerEvents !== 'none';
            """,
            element,
        )
    except WebDriverException:
        script_enabled = True

    return bool(enabled and script_enabled)


def get_special_answer_phrases(question_text: str) -> tuple[str, ...]:
    normalized_question = normalize_key(question_text)
    for question_phrases, answer_phrases in SURVEY_SPECIAL_QUESTION_RULES:
        if any(phrase in normalized_question for phrase in question_phrases):
            return answer_phrases
    return ()


def choose_matching_option(question_text: str, options: list[SurveyOption]) -> SurveyOption | None:
    if not options:
        return None

    answer_phrases = get_special_answer_phrases(question_text)
    for phrase in answer_phrases:
        matches = [option for option in options if text_contains_phrase(option.normalized_text, phrase)]
        if matches:
            return random.choice(matches)

    for phrase in SURVEY_DEFAULT_OPTION_PRIORITY:
        matches = [option for option in options if text_contains_phrase(option.normalized_text, phrase)]
        if matches:
            return random.choice(matches)

    return random.choice(options)


def build_survey_option_groups(options: list[SurveyOption]) -> list[list[SurveyOption]]:
    if not options:
        return []

    if all(option.kind == "button" for option in options):
        return [options]

    grouped_options: dict[tuple[str, str], list[SurveyOption]] = {}
    for option in options:
        if option.kind == "button":
            continue
        group_key = option.group_key or option.normalized_text
        grouped_options.setdefault((option.kind, group_key), []).append(option)

    if grouped_options:
        return list(grouped_options.values())
    return [options]


def set_single_choice_selection(
    driver: webdriver.Chrome,
    selected_option: SurveyOption,
    group_options: list[SurveyOption],
) -> None:
    if selected_option.control is None:
        raise RuntimeError("Selected survey option is missing the input element.")

    for option in group_options:
        if option.control is None:
            continue
        try:
            is_selected = option.control.is_selected()
        except WebDriverException:
            is_selected = False

        should_be_selected = option.control == selected_option.control
        if option.kind == "radio":
            if should_be_selected and not is_selected:
                click_element(driver, option.element)
            continue
        if should_be_selected == is_selected:
            continue
        click_element(driver, option.element)


def click_action_by_keywords(driver: webdriver.Chrome, keywords: tuple[str, ...]) -> str | None:
    actions = collect_clickable_actions(driver)
    match = find_action_by_keywords(actions, keywords)
    if match is None:
        return None
    element, text, _ = match
    click_element(driver, element)
    return text


def has_action_by_keywords(actions: list[tuple[Any, str, str]], keywords: tuple[str, ...]) -> bool:
    return find_action_by_keywords(actions, keywords) is not None


def page_state_changed(
    driver: webdriver.Chrome,
    previous_url: str,
    previous_text: str,
    previous_handles: tuple[str, ...],
) -> bool:
    try:
        current_handles = tuple(driver.window_handles)
        current_url = driver.current_url
        current_text = get_page_text(driver)
    except WebDriverException:
        return True
    return current_handles != previous_handles or current_url != previous_url or current_text != previous_text


def wait_for_page_change(
    driver: webdriver.Chrome,
    previous_url: str,
    previous_text: str,
    previous_handles: tuple[str, ...],
    timeout_seconds: float,
) -> bool:
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        try:
            current_handles = tuple(driver.window_handles)
            current_url = driver.current_url
            current_text = get_page_text(driver)
        except WebDriverException:
            return True
        if current_handles != previous_handles or current_url != previous_url or current_text != previous_text:
            return True
        time.sleep(0.25)
    return False


def retry_survey_page(driver: webdriver.Chrome) -> None:
    time.sleep(random.uniform(5, 7))
    driver.refresh()


def page_looks_complete(page_text: str) -> bool:
    normalized_page_text = normalize_key(page_text)
    return any(phrase in normalized_page_text for phrase in SURVEY_COMPLETION_PAGE_KEYWORDS)


def is_offer_wall_page(page_text: str, actions: list[tuple[Any, str, str]]) -> bool:
    normalized_page_text = normalize_key(page_text)
    if any(phrase in normalized_page_text for phrase in SURVEY_OFFER_PAGE_KEYWORDS):
        return True
    offer_like_actions = 0
    for _, _, normalized_text in actions:
        if any(phrase in normalized_text for phrase in SURVEY_ACTION_EXCLUDE_PHRASES):
            continue
        if any(
            normalized_text == keyword
            or normalized_text.startswith(f"{keyword} ")
            or text_contains_phrase(normalized_text, keyword)
            for keyword in SURVEY_OFFER_ACTION_PRIORITY
        ):
            offer_like_actions += 1
    return offer_like_actions >= 2


def click_offer_wall_action(driver: webdriver.Chrome) -> str | None:
    actions = collect_clickable_actions(driver)

    filtered_actions: list[tuple[Any, str, str]] = []
    for element, text, normalized_text in actions:
        if any(phrase in normalized_text for phrase in SURVEY_ACTION_EXCLUDE_PHRASES):
            continue
        filtered_actions.append((element, text, normalized_text))

    for keyword in SURVEY_OFFER_ACTION_PRIORITY:
        for element, text, normalized_text in filtered_actions:
            if normalized_text == keyword or normalized_text.startswith(f"{keyword} "):
                click_element(driver, element)
                return text

    for keyword in SURVEY_OFFER_ACTION_PRIORITY:
        for element, text, normalized_text in filtered_actions:
            if text_contains_phrase(normalized_text, keyword):
                click_element(driver, element)
                return text

    return None


def inspect_current_survey_context(driver: webdriver.Chrome) -> SurveyPageContext:
    page_text = get_page_text(driver)
    radio_options = collect_radio_options(driver)
    checkbox_options = collect_checkbox_options(driver)
    button_options = collect_button_options(driver)
    actions = collect_clickable_actions(driver)
    return SurveyPageContext(
        handle=driver.current_window_handle,
        url=driver.current_url,
        page_text=page_text,
        radio_options=radio_options,
        checkbox_options=checkbox_options,
        button_options=button_options,
        actions=actions,
    )


def score_survey_context(
    context: SurveyPageContext,
    submit_reference: tuple[str, str, tuple[str, ...]] | None,
) -> int:
    score = 0
    if context.radio_options or context.checkbox_options:
        score += 100
    elif context.button_options:
        score += 80

    if page_looks_complete(context.page_text):
        score += 60
    if has_action_by_keywords(context.actions, SURVEY_CONTINUE_KEYWORDS):
        score += 40
    if has_action_by_keywords(context.actions, SURVEY_FINAL_CTA_KEYWORDS):
        score += 30

    if submit_reference is not None:
        submit_url, submit_text, submit_handles = submit_reference
        if (
            context.handle not in submit_handles
            or context.url != submit_url
            or context.page_text != submit_text
        ):
            score += 10

    return score


def switch_to_best_survey_context(
    driver: webdriver.Chrome,
    submit_reference: tuple[str, str, tuple[str, ...]] | None = None,
) -> SurveyPageContext:
    try:
        current_handle = driver.current_window_handle
        handles = list(driver.window_handles)
    except WebDriverException:
        return inspect_current_survey_context(driver)

    ordered_handles: list[str] = []
    if current_handle in handles:
        ordered_handles.append(current_handle)

    allowed_handles = list(handles)
    if submit_reference is not None:
        _, _, submit_handles = submit_reference
        new_handles = [handle for handle in handles if handle not in submit_handles]
        allowed_handles = [current_handle, *new_handles]

    for handle in reversed(allowed_handles):
        if handle in handles and handle not in ordered_handles:
            ordered_handles.append(handle)

    best_context: SurveyPageContext | None = None
    best_score = -1

    for handle in ordered_handles:
        try:
            driver.switch_to.window(handle)
            context = inspect_current_survey_context(driver)
        except WebDriverException:
            continue

        score = score_survey_context(context, submit_reference)
        if score > best_score:
            best_context = context
            best_score = score

    if best_context is None:
        driver.switch_to.window(current_handle)
        return inspect_current_survey_context(driver)

    if driver.current_window_handle != best_context.handle:
        driver.switch_to.window(best_context.handle)
        best_context = inspect_current_survey_context(driver)

    return best_context


def handle_post_submit_surveys(
    driver: webdriver.Chrome,
    form_config: dict[str, Any],
    timeout_seconds: int,
    submit_reference: tuple[str, str, tuple[str, ...]] | None = None,
    on_submit_completed: Callable[[str], None] | None = None,
) -> str:
    submit_marked_done = False

    def mark_submit_completed(message: str) -> None:
        nonlocal submit_marked_done
        if submit_marked_done or on_submit_completed is None:
            return
        on_submit_completed(message)
        submit_marked_done = True

    if not bool(form_config.get("handle_surveys_after_submit", True)):
        time.sleep(float(form_config.get("post_submit_wait_seconds", 3)))
        mark_submit_completed("submitted; follow-up survey handling disabled")
        return "submitted"

    time.sleep(float(form_config.get("post_submit_wait_seconds", 3)))
    max_pages = int(form_config.get("survey_max_pages", 12))
    max_retries = int(form_config.get("survey_retry_count", 2))
    settle_wait = float(form_config.get("survey_step_wait_seconds", 1.5))
    pages_completed = 0
    reload_attempts = 0

    while pages_completed < max_pages:
        time.sleep(settle_wait)

        context = switch_to_best_survey_context(driver, submit_reference)
        page_text = context.page_text
        actions = context.actions
        survey_options = context.survey_options

        page_changed_after_submit = True
        if submit_reference is not None:
            submit_url, submit_text, submit_handles = submit_reference
            page_changed_after_submit = page_state_changed(driver, submit_url, submit_text, submit_handles)

        if not submit_marked_done:
            if survey_options or page_looks_complete(page_text):
                mark_submit_completed("submitted; survey page reached")
            elif page_changed_after_submit and (
                has_action_by_keywords(actions, SURVEY_CONTINUE_KEYWORDS)
                or has_action_by_keywords(actions, SURVEY_FINAL_CTA_KEYWORDS)
            ):
                mark_submit_completed("submitted; post-submit page reached")

        if is_offer_wall_page(page_text, actions):
            previous_url = driver.current_url
            previous_text = page_text
            previous_handles = tuple(driver.window_handles)
            offer_text = click_offer_wall_action(driver)
            if offer_text:
                advanced = wait_for_page_change(
                    driver,
                    previous_url,
                    previous_text,
                    previous_handles,
                    timeout_seconds,
                )
                if advanced:
                    time.sleep(2)
                    return f"submitted; offer clicked ({offer_text})"

            if reload_attempts < max_retries:
                reload_attempts += 1
                retry_survey_page(driver)
                continue
            raise RuntimeError("Offer wall page detected but no clickable blue offer button advanced the flow.")

        if not survey_options:
            if page_looks_complete(page_text):
                mark_submit_completed("submitted; completion page detected")
                if pages_completed == 0:
                    return "submitted; completion page detected"
                return f"submitted; survey handled across {pages_completed} page(s)"

            has_continue_action = has_action_by_keywords(actions, SURVEY_CONTINUE_KEYWORDS)
            has_final_cta_action = has_action_by_keywords(actions, SURVEY_FINAL_CTA_KEYWORDS)

            if not has_final_cta_action and reload_attempts < max_retries:
                reload_attempts += 1
                retry_survey_page(driver)
                continue

            previous_url = driver.current_url
            previous_handles = tuple(driver.window_handles)
            continue_text = None
            if has_continue_action:
                continue_text = click_action_by_keywords(driver, SURVEY_CONTINUE_KEYWORDS)
            if continue_text:
                advanced = wait_for_page_change(driver, previous_url, page_text, previous_handles, timeout_seconds)
                if advanced:
                    pages_completed += 1
                    reload_attempts = 0
                    continue

            cta_text = None
            if has_final_cta_action:
                cta_text = click_action_by_keywords(driver, SURVEY_FINAL_CTA_KEYWORDS)
            if cta_text:
                time.sleep(2)
                return f"submitted; final CTA clicked ({cta_text})"

            if reload_attempts < max_retries:
                reload_attempts += 1
                retry_survey_page(driver)
                continue

            if pages_completed == 0:
                return "submitted; no survey detected"
            return f"submitted; survey handled across {pages_completed} page(s)"

        selected_groups: list[tuple[SurveyOption, list[SurveyOption]]] = []
        for option_group in build_survey_option_groups(survey_options):
            question_text = next(
                (option.question_text for option in option_group if normalize_key(option.question_text)),
                page_text,
            )
            selected_option = choose_matching_option(question_text, option_group)
            if selected_option is None:
                if reload_attempts < max_retries:
                    reload_attempts += 1
                    retry_survey_page(driver)
                    break
                raise RuntimeError(f"Could not match a survey option on page text: {page_text[:200]}")
            selected_groups.append((selected_option, option_group))
        else:
            reload_attempts = 0
            previous_url = driver.current_url
            previous_text = page_text
            previous_handles = tuple(driver.window_handles)

            for selected_option, option_group in selected_groups:
                if selected_option.kind in {"checkbox", "radio"}:
                    set_single_choice_selection(driver, selected_option, option_group)
                else:
                    click_element(driver, selected_option.element)
                    break

            advanced = wait_for_page_change(driver, previous_url, previous_text, previous_handles, 2.5)
            if not advanced:
                continue_text = click_action_by_keywords(driver, SURVEY_CONTINUE_KEYWORDS)
                if continue_text:
                    advanced = wait_for_page_change(
                        driver,
                        previous_url,
                        previous_text,
                        previous_handles,
                        timeout_seconds,
                    )

            if not advanced:
                cta_text = click_action_by_keywords(driver, SURVEY_FINAL_CTA_KEYWORDS)
                if cta_text:
                    time.sleep(2)
                    return f"submitted; final CTA clicked ({cta_text})"

                if reload_attempts < max_retries:
                    reload_attempts += 1
                    retry_survey_page(driver)
                    continue
                selected_text = ", ".join(selected_option.text for selected_option, _ in selected_groups)
                raise RuntimeError(f"Survey answer '{selected_text}' did not advance the page.")

            pages_completed += 1
            continue

        continue

    raise RuntimeError(f"Survey flow exceeded the configured limit of {max_pages} page(s).")


def set_input_value(driver: webdriver.Chrome, element, value: str) -> None:
    driver.execute_script(
        """
        const element = arguments[0];
        const value = arguments[1];
        element.focus();
        element.value = value;
        element.dispatchEvent(new Event('input', { bubbles: true }));
        element.dispatchEvent(new Event('change', { bubbles: true }));
        """,
        element,
        value,
    )


def clear_and_type(driver: webdriver.Chrome, element, value: str) -> None:
    try:
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
        element.click()
        element.send_keys(Keys.CONTROL, "a")
        element.send_keys(Keys.DELETE)
        if value:
            element.send_keys(value)
        return
    except WebDriverException:
        set_input_value(driver, element, value)


def set_checkbox_state(driver: webdriver.Chrome, selectors: list[str], desired_state: bool, timeout_seconds: int) -> None:
    element = find_first_element(driver, selectors, timeout_seconds, require_displayed=False)
    try:
        is_checked = element.is_selected()
    except WebDriverException:
        is_checked = False
    if is_checked != desired_state:
        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
            element.click()
        except WebDriverException:
            driver.execute_script(
                """
                arguments[0].checked = arguments[1];
                arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                """,
                element,
                desired_state,
            )


def fill_field(driver: webdriver.Chrome, row_values: dict[str, str], field_config: dict[str, Any], timeout_seconds: int) -> None:
    column_name = field_config["column"]
    field_type = field_config.get("type", "text")
    selectors = get_field_selectors(field_config)
    if not selectors:
        raise ValueError(f"No selectors configured for Excel column '{column_name}'")

    raw_value = get_field_value(row_values, field_config)
    value = transform_value(raw_value, field_config)

    if field_type == "text":
        element = find_first_element(driver, selectors, timeout_seconds, require_displayed=True)
        clear_and_type(driver, element, value)
        return

    if field_type == "select_text":
        element = find_first_element(driver, selectors, timeout_seconds, require_displayed=True)
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
        if not value:
            return
        from selenium.webdriver.support.ui import Select

        select = Select(element)
        try:
            select.select_by_visible_text(value)
        except Exception:
            try:
                select.select_by_value(value)
            except Exception:
                driver.execute_script(
                    """
                    arguments[0].value = arguments[1];
                    arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                    arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    """,
                    element,
                    value,
                )
        return

    if field_type == "checkbox":
        desired_state = value.lower() in {"1", "true", "yes", "y"}
        set_checkbox_state(driver, selectors, desired_state, timeout_seconds)
        return

    if field_type == "hidden_text":
        element = find_first_element(driver, selectors, timeout_seconds, require_displayed=False)
        driver.execute_script("arguments[0].value = arguments[1];", element, value)
        return

    raise ValueError(f"Unsupported field type '{field_type}' for column '{column_name}'")


def apply_static_checkboxes(driver: webdriver.Chrome, form_config: dict[str, Any], timeout_seconds: int) -> None:
    for checkbox in form_config.get("checkboxes", []):
        selectors = checkbox.get("selectors") or ([checkbox["selector"]] if checkbox.get("selector") else [])
        if not selectors:
            continue
        desired_state = bool(checkbox.get("checked", True))
        set_checkbox_state(driver, selectors, desired_state, timeout_seconds)


def get_initial_step_field_indexes(fields: list[dict[str, Any]], form_config: dict[str, Any]) -> list[int]:
    zip_step_config = form_config.get("zip_step") or {}
    configured_columns = zip_step_config.get("field_columns") or FORM_INITIAL_STEP_DEFAULT_COLUMNS
    normalized_columns = {normalize_key(value) for value in configured_columns if str(value).strip()}
    if not normalized_columns:
        return []

    indexes: list[int] = []
    for index, field_config in enumerate(fields):
        column_name = str(field_config.get("column", "")).strip()
        if column_name and normalize_key(column_name) in normalized_columns:
            indexes.append(index)
    return indexes


def find_initial_step_action(driver: webdriver.Chrome, form_config: dict[str, Any]):
    zip_step_config = form_config.get("zip_step") or {}
    next_selectors = [str(item).strip() for item in zip_step_config.get("next_selectors") or [] if str(item).strip()]
    if next_selectors:
        button = try_find_first_element(driver, next_selectors, require_displayed=True)
        if button is None:
            return None
        label = get_element_text(driver, button) or next_selectors[0]
        return button, label, normalize_key(label)

    next_keywords = tuple(
        str(item).strip().lower()
        for item in zip_step_config.get("next_keywords") or FORM_INITIAL_STEP_DEFAULT_ACTION_KEYWORDS
        if str(item).strip()
    )
    return find_action_by_keywords(collect_clickable_actions(driver), next_keywords)


def wait_for_initial_step_action(
    driver: webdriver.Chrome,
    form_config: dict[str, Any],
    timeout_seconds: float,
):
    deadline = time.time() + max(timeout_seconds, 0)
    while True:
        action = find_initial_step_action(driver, form_config)
        if action is not None:
            element, text, normalized_text = action
            if element_is_enabled(driver, element):
                return element, text, normalized_text
        if time.time() >= deadline:
            break
        time.sleep(0.25)
    return None


def click_initial_step_action(driver: webdriver.Chrome, form_config: dict[str, Any], timeout_seconds: int) -> str:
    action = wait_for_initial_step_action(driver, form_config, timeout_seconds)
    if action is not None:
        element, text, _ = action
        click_element(driver, element)
        return text

    raise RuntimeError(
        "ZIP step was detected but no enabled Next/Continue button was found. "
        "Set form.zip_step.next_selectors in config.json if needed."
    )


def initial_step_action_is_available(driver: webdriver.Chrome, form_config: dict[str, Any]) -> bool:
    action = find_initial_step_action(driver, form_config)
    if action is None:
        return False
    element, _, _ = action
    return element_is_enabled(driver, element)


def initial_step_advanced(
    driver: webdriver.Chrome,
    previous_url: str,
    previous_text: str,
    previous_handles: tuple[str, ...],
    follow_up_fields: list[dict[str, Any]],
    transition_wait: float,
) -> bool:
    page_changed = wait_for_page_change(
        driver,
        previous_url,
        previous_text,
        previous_handles,
        transition_wait,
    )
    if page_changed:
        return True
    if follow_up_fields and wait_for_any_field_visible(driver, follow_up_fields, transition_wait):
        return True
    return False


def maybe_complete_initial_zip_step(
    driver: webdriver.Chrome,
    row_values: dict[str, str],
    form_config: dict[str, Any],
    fields: list[dict[str, Any]],
    timeout_seconds: int,
) -> set[int]:
    zip_step_config = form_config.get("zip_step") or {}
    if not bool(zip_step_config.get("enabled", True)):
        return set()

    initial_step_indexes = get_initial_step_field_indexes(fields, form_config)
    if not initial_step_indexes:
        return set()

    visible_initial_step_indexes = [
        index for index in initial_step_indexes if field_is_visible_now(driver, fields[index])
    ]
    if not visible_initial_step_indexes:
        return set()

    follow_up_fields = [
        field_config
        for index, field_config in enumerate(fields)
        if index not in initial_step_indexes and str(field_config.get("type", "text")).lower() != "hidden_text"
    ]
    if follow_up_fields and any(field_is_visible_now(driver, field_config) for field_config in follow_up_fields):
        return set()

    detection_wait = float(zip_step_config.get("detection_wait_seconds", 1.5))
    if follow_up_fields and detection_wait > 0 and wait_for_any_field_visible(driver, follow_up_fields, detection_wait):
        return set()
    if not initial_step_action_is_available(driver, form_config):
        return set()

    zip_step_config = form_config.get("zip_step") or {}
    for index in visible_initial_step_indexes:
        fill_field(driver, row_values, fields[index], timeout_seconds)
        filled_element = get_visible_field_element(driver, fields[index])
        if filled_element is not None:
            blur_element(driver, filled_element)

    after_fill_wait = float(zip_step_config.get("after_fill_wait_seconds", 0.75))
    if after_fill_wait > 0:
        time.sleep(after_fill_wait)

    action_ready_wait = float(zip_step_config.get("action_ready_wait_seconds", 5))
    advance_retry_count = max(1, int(zip_step_config.get("advance_retry_count", 3)))
    retry_pause_seconds = float(zip_step_config.get("retry_pause_seconds", 1.0))
    transition_wait = float(zip_step_config.get("transition_wait_seconds", timeout_seconds))

    last_action_text = ""
    for attempt in range(1, advance_retry_count + 1):
        previous_url = driver.current_url
        previous_text = get_page_text(driver)
        previous_handles = tuple(driver.window_handles)

        action = wait_for_initial_step_action(driver, form_config, action_ready_wait)
        if action is None:
            raise RuntimeError(
                "ZIP step was detected but the Next/Continue button did not become ready. "
                "Set form.zip_step.next_selectors in config.json if needed."
            )

        element, action_text, _ = action
        last_action_text = action_text
        log(f"ZIP step attempt {attempt}/{advance_retry_count} using action '{action_text}'")
        click_element(driver, element)

        if initial_step_advanced(
            driver,
            previous_url,
            previous_text,
            previous_handles,
            follow_up_fields,
            transition_wait,
        ):
            log(f"ZIP step completed with action '{action_text}'")
            return set(initial_step_indexes)

        if attempt < advance_retry_count and retry_pause_seconds > 0:
            log(f"ZIP step did not advance on attempt {attempt}. Retrying after {retry_pause_seconds} second(s).")
            time.sleep(retry_pause_seconds)

    raise RuntimeError(
        f"ZIP step clicked '{last_action_text or 'Next/Continue'}' but the next form did not appear "
        f"within {transition_wait} second(s) after {advance_retry_count} attempt(s)."
    )


def submit_form(driver: webdriver.Chrome, form_config: dict[str, Any], timeout_seconds: int) -> None:
    submit_selector = str(form_config.get("submit_selector", "")).strip()
    if not submit_selector:
        raise ValueError("submit_after_fill is enabled but form.submit_selector is empty")

    button = find_first_element(driver, [submit_selector], timeout_seconds)
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
    try:
        button.click()
    except WebDriverException:
        driver.execute_script("arguments[0].click();", button)

    success_wait_selector = str(form_config.get("success_wait_selector", "")).strip()
    if success_wait_selector:
        find_first_element(driver, [success_wait_selector], timeout_seconds)
        return


def remember_form_location(
    driver: webdriver.Chrome,
    form_config: dict[str, Any],
    runtime_state: FormRuntimeState | None,
    force: bool = False,
) -> None:
    if runtime_state is None:
        return
    target_contains = str(form_config.get("target_url_contains", "")).strip().lower()
    current_url = driver.current_url
    if not current_url:
        return
    if not force and target_contains and target_contains not in current_url.lower():
        return
    runtime_state.last_form_url = current_url
    runtime_state.last_form_handle = driver.current_window_handle


def page_matches_form_target(driver: webdriver.Chrome, form_config: dict[str, Any]) -> bool:
    target_contains = str(form_config.get("target_url_contains", "")).strip().lower()
    try:
        current_url = driver.current_url.lower()
    except WebDriverException:
        return False

    if target_contains and target_contains in current_url:
        return True

    submit_selector = str(form_config.get("submit_selector", "")).strip()
    if not submit_selector:
        return False

    try:
        by, query = parse_selector(submit_selector)
        for element in driver.find_elements(by, query):
            try:
                if element.is_displayed():
                    return True
            except WebDriverException:
                return True
    except WebDriverException:
        return False

    fields = form_config.get("__resolved_fields__") or []
    if fields and page_has_configured_field(driver, fields):
        return True

    return False


def extract_host_hint(target_contains: str) -> str:
    match = re.search(r"([a-z0-9.-]+\.[a-z]{2,})", str(target_contains).lower())
    if not match:
        return ""
    return match.group(1)


def try_recover_form_with_history(
    driver: webdriver.Chrome,
    form_config: dict[str, Any],
    runtime_state: FormRuntimeState | None = None,
) -> bool:
    max_steps = max(0, int(form_config.get("history_recover_max_steps", 3)))
    if max_steps <= 0:
        return False

    host_hint = extract_host_hint(str(form_config.get("target_url_contains", "")).strip())
    if not host_hint:
        return False

    try:
        handles = list(driver.window_handles)
    except WebDriverException:
        return False

    for handle in reversed(handles):
        try:
            driver.switch_to.window(handle)
            current_url = driver.current_url.lower()
        except WebDriverException:
            continue

        if host_hint not in current_url:
            continue

        for _ in range(max_steps):
            if page_matches_form_target(driver, form_config):
                remember_form_location(driver, form_config, runtime_state, force=True)
                return True

            previous_url = driver.current_url
            previous_text = get_page_text(driver)
            previous_handles = tuple(driver.window_handles)
            try:
                driver.back()
            except WebDriverException:
                break

            wait_for_page_change(driver, previous_url, previous_text, previous_handles, 2.5)
            if page_matches_form_target(driver, form_config):
                remember_form_location(driver, form_config, runtime_state, force=True)
                return True

            try:
                if driver.current_url == previous_url:
                    break
            except WebDriverException:
                break

    return False


def focus_browser_tab(
    driver: webdriver.Chrome,
    form_config: dict[str, Any],
    runtime_state: FormRuntimeState | None = None,
) -> bool:
    handles = driver.window_handles
    if not handles:
        raise RuntimeError("No browser tabs are open in this AdsPower profile.")

    target_contains = str(form_config.get("target_url_contains", "")).strip().lower()

    if runtime_state and runtime_state.last_form_handle and runtime_state.last_form_handle in handles:
        driver.switch_to.window(runtime_state.last_form_handle)
        if page_matches_form_target(driver, form_config):
            remember_form_location(driver, form_config, runtime_state, force=True)
            return True

    current_handle = driver.current_window_handle
    if current_handle in handles and page_matches_form_target(driver, form_config):
        remember_form_location(driver, form_config, runtime_state, force=True)
        return True

    for handle in reversed(handles):
        driver.switch_to.window(handle)
        if page_matches_form_target(driver, form_config):
            remember_form_location(driver, form_config, runtime_state, force=True)
            return True

    if try_recover_form_with_history(driver, form_config, runtime_state):
        return True

    if runtime_state and runtime_state.last_form_handle and runtime_state.last_form_handle in handles:
        driver.switch_to.window(runtime_state.last_form_handle)
        if runtime_state.last_form_url:
            return False
        if not target_contains:
            remember_form_location(driver, form_config, runtime_state, force=True)
            return True

    if current_handle in handles:
        driver.switch_to.window(current_handle)
        if not target_contains:
            remember_form_location(driver, form_config, runtime_state, force=True)
            return True

    tab_index = int(form_config.get("tab_index", -1))
    try:
        target_handle = handles[tab_index]
    except IndexError as exc:
        raise ValueError(
            f"Configured form.tab_index={tab_index} but only {len(handles)} tab(s) are open in this profile."
        ) from exc

    driver.switch_to.window(target_handle)
    if page_matches_form_target(driver, form_config):
        remember_form_location(driver, form_config, runtime_state, force=True)
        return True
    if runtime_state and runtime_state.last_form_url:
        return False
    remember_form_location(driver, form_config, runtime_state, force=True)
    return not target_contains or target_contains in driver.current_url.lower()


def open_form(
    driver: webdriver.Chrome,
    form_config: dict[str, Any],
    runtime_state: FormRuntimeState | None = None,
) -> None:
    if form_config.get("use_existing_page", False):
        found_target = focus_browser_tab(driver, form_config, runtime_state)
        if not found_target and runtime_state and runtime_state.last_form_url:
            page_load_timeout = int(form_config.get("page_load_timeout_seconds", 30))
            driver.set_page_load_timeout(page_load_timeout)
            driver.get(runtime_state.last_form_url)
            remember_form_location(driver, form_config, runtime_state)
        return

    target_url = str(form_config["url"]).strip()
    if not target_url:
        raise ValueError("form.url is empty in config and form.use_existing_page is false")
    page_load_timeout = int(form_config.get("page_load_timeout_seconds", 30))
    driver.set_page_load_timeout(page_load_timeout)
    driver.get(target_url)
    remember_form_location(driver, form_config, runtime_state)


def process_row(
    driver: webdriver.Chrome,
    row_task: RowTask,
    config: dict[str, Any],
    runtime_state: FormRuntimeState | None = None,
    on_submit_completed: Callable[[str], None] | None = None,
) -> str:
    form_config = config["form"]
    fields = config["fields"]
    timeout_seconds = int(form_config.get("field_timeout_seconds", 15))

    if form_config.get("return_to_form_before_each_row", True):
        open_form(driver, form_config, runtime_state)
        remember_form_location(driver, form_config, runtime_state, force=True)

    skip_field_indexes = maybe_complete_initial_zip_step(
        driver,
        row_task.values,
        form_config,
        fields,
        timeout_seconds,
    )

    for index, field_config in enumerate(fields):
        if index in skip_field_indexes:
            continue
        fill_field(driver, row_task.values, field_config, timeout_seconds)

    apply_static_checkboxes(driver, form_config, timeout_seconds)

    if form_config.get("submit_after_fill", False):
        submit_reference = (
            driver.current_url,
            get_page_text(driver),
            tuple(driver.window_handles),
        )
        submit_form(driver, form_config, timeout_seconds)
        return handle_post_submit_surveys(
            driver,
            form_config,
            timeout_seconds,
            submit_reference=submit_reference,
            on_submit_completed=on_submit_completed,
        )

    return "filled (submit disabled)"


def worker_loop(profile_target: dict[str, Any], assigned_tasks: list[RowTask], tracker: ExcelTracker, config: dict[str, Any]) -> None:
    client = AdsPowerClient(config["adspower"])
    retry_count = int(config.get("worker", {}).get("retry_count", 1))
    max_rows_per_profile = int(config.get("worker", {}).get("max_rows_per_profile", 0))
    profile_id = str(profile_target["user_id"])
    connect_mode = str(profile_target.get("connect_mode", "start"))

    driver = None
    processed_count = 0
    runtime_state = FormRuntimeState()

    try:
        if connect_mode == "attach":
            log(f"[{profile_id}] Attaching to already open AdsPower browser")
            driver = client.attach_to_active_profile(profile_target["browser_data"])
        else:
            log(f"[{profile_id}] Starting AdsPower profile")
            driver, _ = client.start_profile(profile_id)

        if config["form"].get("return_to_form_before_each_row", True):
            open_form(driver, config["form"], runtime_state)

        for row_task in assigned_tasks:
            if max_rows_per_profile and processed_count >= max_rows_per_profile:
                log(f"[{profile_id}] Reached max_rows_per_profile={max_rows_per_profile}")
                break
            result_status = "FAILED"
            result_message = ""
            done_marked_early = False

            def mark_done_after_submit(message: str) -> None:
                nonlocal done_marked_early
                tracker.mark_result(row_task.row_number, "DONE", message, profile_id)
                done_marked_early = True

            for attempt in range(1, retry_count + 2):
                try:
                    log(f"[{profile_id}] Processing Excel row {row_task.row_number} (attempt {attempt})")
                    result_message = process_row(
                        driver,
                        row_task,
                        config,
                        runtime_state=runtime_state,
                        on_submit_completed=mark_done_after_submit,
                    )
                    result_status = "DONE"
                    break
                except Exception as exc:
                    result_message = str(exc)
                    if done_marked_early:
                        result_status = "DONE"
                        result_message = f"main form submitted; survey follow-up stopped after DONE: {exc}"
                        log(
                            f"[{profile_id}] Row {row_task.row_number} already marked DONE after submit. "
                            f"Stopping follow-up retries: {exc}"
                        )
                        break
                    if attempt <= retry_count:
                        log(
                            f"[{profile_id}] Row {row_task.row_number} failed on attempt {attempt}: {exc}. Retrying..."
                        )
                        time.sleep(2)
                    else:
                        log(f"[{profile_id}] Row {row_task.row_number} failed: {exc}")

            tracker.mark_result(row_task.row_number, result_status, result_message, profile_id)
            processed_count += 1
            if "final cta clicked" in result_message.lower() or "offer clicked" in result_message.lower():
                log(f"[{profile_id}] Final offer page reached, stopping this browser worker.")
                break

    except Exception as exc:
        log(f"[{profile_id}] Worker stopped: {exc}")
    finally:
        if driver is not None and connect_mode == "start" and not client.keep_browser_open:
            try:
                driver.quit()
            except Exception:
                pass
        if connect_mode == "start":
            try:
                client.stop_profile(profile_id)
            except Exception as exc:
                log(f"[{profile_id}] Could not stop AdsPower profile cleanly: {exc}")


def validate_config(config: dict[str, Any], tracker: ExcelTracker) -> None:
    if not config.get("fields"):
        raise ValueError("No fields configured in config.json")

    profile_ids = config.get("adspower", {}).get("profile_ids") or []
    use_active_profiles = bool(config.get("adspower", {}).get("use_active_profiles", True))
    if not profile_ids and not use_active_profiles:
        raise ValueError("adspower.profile_ids is empty")

    tracker.assert_field_sources_exist(config["fields"])

    transforms = {
        str(field.get("transform", "")).strip().lower()
        for field in config["fields"]
        if str(field.get("transform", "")).strip()
    }
    supported_transforms = {"date_month_number", "date_day_number", "date_year"}
    unsupported = transforms - supported_transforms
    if unsupported:
        raise ValueError(f"Unsupported transforms in config: {', '.join(sorted(unsupported))}")

    zip_step_config = config.get("form", {}).get("zip_step") or {}
    configured_columns = zip_step_config.get("field_columns") or FORM_INITIAL_STEP_DEFAULT_COLUMNS
    next_keywords = tuple(
        str(item).strip()
        for item in zip_step_config.get("next_keywords") or FORM_INITIAL_STEP_DEFAULT_ACTION_KEYWORDS
        if str(item).strip()
    )
    next_selectors = [str(item).strip() for item in zip_step_config.get("next_selectors") or [] if str(item).strip()]
    if zip_step_config.get("enabled", True) and configured_columns and not next_keywords and not next_selectors:
        raise ValueError(
            "form.zip_step is enabled but no next_keywords or next_selectors are configured."
        )


def resolve_profile_targets(config: dict[str, Any]) -> list[dict[str, Any]]:
    client = AdsPowerClient(config["adspower"])
    explicit_profile_ids = [str(item).strip() for item in config["adspower"].get("profile_ids", []) if str(item).strip()]
    if explicit_profile_ids:
        return [{"user_id": profile_id, "connect_mode": "start"} for profile_id in explicit_profile_ids]

    if bool(config["adspower"].get("use_active_profiles", True)):
        active_profiles = client.list_active_profiles()
        if not active_profiles:
            raise RuntimeError(
                "No open AdsPower browsers found. Open the desired AdsPower browser profiles first, then run again."
            )
        return [
            {
                "user_id": str(profile_data["user_id"]),
                "connect_mode": "attach",
                "browser_data": profile_data,
            }
            for profile_data in active_profiles
        ]

    raise RuntimeError("No AdsPower profiles available to run.")


def build_profile_assignments(
    selected_profile_targets: list[dict[str, Any]],
    selected_tasks: list[RowTask],
    max_rows_per_profile: int,
) -> list[tuple[dict[str, Any], list[RowTask]]]:
    if not selected_profile_targets or not selected_tasks:
        return []

    if max_rows_per_profile <= 0:
        assignments = [(profile_target, []) for profile_target in selected_profile_targets]
        for index, row_task in enumerate(selected_tasks):
            assignments[index % len(assignments)][1].append(row_task)
        return [(profile_target, assigned) for profile_target, assigned in assignments if assigned]

    assignments: list[tuple[dict[str, Any], list[RowTask]]] = []
    start_index = 0
    rows_per_profile = max_rows_per_profile
    for profile_target in selected_profile_targets:
        assigned = selected_tasks[start_index : start_index + rows_per_profile]
        if not assigned:
            break
        assignments.append((profile_target, assigned))
        start_index += rows_per_profile
    return assignments


def run(config_path: Path) -> None:
    config = load_config(config_path)
    config["form"]["__resolved_fields__"] = config.get("fields", [])
    tracker = ExcelTracker(config["excel"])
    validate_config(config, tracker)
    profile_targets = resolve_profile_targets(config)

    pending_tasks = tracker.build_pending_tasks()
    if not pending_tasks:
        log("No pending Excel rows found. Rows marked DONE are skipped automatically.")
        return

    max_rows_per_profile = int(config.get("worker", {}).get("max_rows_per_profile", 0))
    if max_rows_per_profile > 0:
        run_row_cap = len(profile_targets) * max_rows_per_profile
        selected_tasks = pending_tasks[:run_row_cap]
        workers_needed = math.ceil(len(selected_tasks) / max_rows_per_profile) if selected_tasks else 0
        selected_profile_targets = profile_targets[:workers_needed]
    else:
        selected_tasks = pending_tasks
        selected_profile_targets = profile_targets

    assignments = build_profile_assignments(selected_profile_targets, selected_tasks, max_rows_per_profile)

    log(
        f"Found {len(pending_tasks)} pending rows and {len(profile_targets)} AdsPower browser(s). "
        f"This run will process up to {len(selected_tasks)} row(s)."
    )

    threads: list[threading.Thread] = []
    for profile_target, assigned_tasks in assignments:
        thread = threading.Thread(
            target=worker_loop,
            args=(profile_target, assigned_tasks, tracker, config),
            daemon=False,
        )
        threads.append(thread)
        thread.start()

    for thread in threads:
        thread.join()

    refreshed_pending = tracker.build_pending_tasks()
    total_remaining = len(refreshed_pending)
    processed_in_batch = len(selected_tasks) - sum(
        1 for task in selected_tasks if any(p.row_number == task.row_number for p in refreshed_pending)
    )
    log(
        f"Batch finished. {processed_in_batch} row(s) were marked DONE in this run, "
        f"and {total_remaining} pending row(s) remain for the next run."
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Fill a web form from Excel rows across multiple AdsPower browsers.")
    parser.add_argument(
        "--config",
        default="config.json",
        help="Path to config JSON file. Default: config.json",
    )
    args = parser.parse_args()
    run(Path(args.config).expanduser().resolve())


if __name__ == "__main__":
    main()
